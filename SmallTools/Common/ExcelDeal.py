#!/usr/bin/env python
# -*- coding: utf-8 -*-

# @Time         = 2017/7/1 9:24
# @Author       = Demon
# @File         = ExcelDeal.py
# @Software     = PyCharm
# @Description  = read and write excel

import xlrd
import xlwt
import os
import Common.LogControl
import logging
import sys
import openpyxl
from datetime import date, datetime


global LOG_FILE, LOG_SECTION, LOG
LOG_FILE = os.getcwd().replace('\\', '/') + '/Common/exceldeal.log'
LOG_SECTION = 'Excel_log'
LOG = Common.LogControl.SetLogger(LOG_FILE, LOG_SECTION)

'''
style_n = easyxf('font: height 200, name 宋体;'
                 'borders: left thin, right thin, top thin, bottom thin;'
                 'pattern: fore_colour balck;'
                 'alignment: horizontal left, vertical center;'
                 num_format_str='#.##0'0
                 )
                
'''


class Excel2003:
    """ 需要xlrd xlwt 模块 """
    xlspath = None
    writewb = None
    readwb = None
    # read = 'r' or write = 'w'
    option = None
    charset = 'utf-8'

    global LOG


    def __init__(self, path, option, charset):
        if not os.path.exists(path) or not os.path.isfile(path):
            LOG.error("[Excel2003 init] file not exist.")
            sys.exit(1)
        LOG.debug("[Excel2003 init] file path: %s." % path)
        self.xlspath = path
        self.charset = charset
        self.option = option
        if 'r' == option:
            self.readwb = xlrd.open_workbook(self.xlspath, encoding_override=self.charset, formatting_info=True)
        elif 'w' == option:
            self.writewb = xlwt.Workbook(encoding=charset)


    def GetWriteWb(self):
        return self.writewb


    def GetReadWb(self):
        return self.readwb


    def GetSheet(self, sheetname):
        sht = None
        if 'r' == self.option:
            if None != sheetname and sheetname in self.readwb.sheet_names():
                sht = self.readwb.sheet_by_name(sheetname)
            else:
                LOG.error("[Excel2003 GetSheet] %s has not sheet named %s." %
                          (os.path.basename(self.xlspath), sheetname))
                return None
        elif 'w' == self.option:
            if None != sheetname or '' == sheetname.strip():
                LOG.warn("[Excel2003 GetSheet] input sheetname is empty, add default sheet1.")
                sheetname = 'sheet1'
                sht = self.writewb.add_sheet(sheetname, cell_overwrite_ok=True)
        return sht


    def CellRead(self, sht, row, col):
        if 'r' != self.option:
            LOG.warn("[Excel2003 CellRead] not read option which is initted.")
            return

        flag = True
        if None == sht or (not isinstance(sht, xlrd.sheet)):
            LOG.error("[Excel2003 CellRead] sheet is None.")
            flag = False

        if (row >= sht.nrows) or (col >= sht.ncols):
            cv = ''
            flag = False
        else:
            ct = sht.cell_type(row, col)
            cv = sht.cell_value(row, col)

            if xlrd.XL_CELL_TEXT == ct or xlrd.XL_CELL_BLANK == ct:
                cv = str(cv).strip()
            elif xlrd.XL_CELL_NUMBER == ct:
                if isinstance(cv, float):
                    cv = str(float(cv))
                elif isinstance(cv, int):
                    cv = str(int(cv))
            elif xlrd.XL_CELL_EMPTY == ct:
                cv = ''
            elif xlrd.XL_CELL_DATE == ct:
                date_value = xlrd.xldate_as_tuple(cv, datemode=0)
                cv = date(date_value[:3]).strftime('%Y/%m/%d')
            elif xlrd.XL_CELL_BOOLEAN == ct:
                if cv:
                    cv = 'True'
                else:
                    cv = 'False'
            else:
                cv = ''

        if flag and None != cv:
            return cv
        else:
            return ''

        pass


    def CellWrite(self, sht, value, startrow=0, startcol=0, endrow=0, endcol=0, style=None):
        if 'w' != self.option:
            LOG.warn("[Excel2003 CellWrite] not write option which is initted.")
            return

        if None == value:
            LOG.error("[Excel2003 CellWrite] input value is empty.")

        if endrow <= startrow:
            endrow = startrow
        if endcol <= startcol:
            endcol = startcol

        if (endrow == startrow) and (endcol == startcol):
            sht.write(startrow, startcol, value, style)
        else:
            sht.write_merge(startrow, endrow, startcol, endcol, value, style)

        LOG.debug("range[(%d, %d)->(%d, %d)] write context %s." % (startrow, startcol, endrow, endcol, value))
        pass


    def SaveWriteWb(self):
        self.writewb.save(self.xlspath)


class Excel2007(Excel2003):
    """ 需要 openpyxls 模块 """
    xlspath = None
    writewb = None
    readwb = None
    # read = 'r' or write = 'w'
    option = None
    charset = 'utf-8'

    global LOG


    def __init__(self, path, option, charset):
        if not os.path.exists(path) or not os.path.isfile(path):
            LOG.error("[Excel2007 init] file not exist.")
            sys.exit(1)
        LOG.debug("[Excel2007 init] file path: %s." % path)
        self.xlspath = path
        self.charset = charset
        self.option = option
        if 'r' == option:
            self.readwb = openpyxl.load_workbook(path)
            self.readwb.encoding = charset
            self.readwb.get_sheet_by_name
        elif 'w' == option:
            self.writewb = openpyxl.Workbook()
            self.writewb.encoding = charset
            self.writewb.create_sheet()


    def GetWriteWb(self):
        return self.writewb


    def GetReadWb(self):
        return self.readwb


    def GetSheet(self, sheetname):
        sht = None
        if 'r' == self.option:
            if None != sheetname and sheetname in self.readwb.get_sheet_names():
                sht = self.readwb.get_sheet_by_name(sheetname)
            else:
                LOG.error("[Excel2007 GetSheet] %s has not sheet named %s." %
                          (os.path.basename(self.xlspath), sheetname))
                return None
        elif 'w' == self.option:
            if None != sheetname or '' == sheetname.strip():
                LOG.warn("[Excel2007 GetSheet] input sheetname is empty, add default sheet1.")
                sheetname = 'sheet1'
                sht = self.writewb.create_sheet(sheetname)
        return sht


    def CellRead(self, sht, row, col):
        if 'r' != self.option:
            LOG.warn("[Excel2007 CellRead] not read option which is initted.")
            return

        flag = True
        if None == sht:
            LOG.error("[Excel2007 CellRead] sheet is None.")
            flag = False

        if (row >= sht.nrows) or (col >= sht.ncols):
            cv = ''
            flag = False
        else:
            cv = sht.cell_value(row, col)

            if isinstance(cv, float):
                cv = str(float(cv))
            elif isinstance(cv, int):
                cv = str(int(cv))
            elif isinstance(cv, bool):
                if cv:
                    cv = 'True'
                else:
                    cv = 'False'
            elif isinstance(cv, str):
                cv = cv.strip()
            elif isinstance(cv, openpyxl.compat.unicode):
                cv = cv.decode('gbk')
            else:
                cv = ''
        return cv

        if flag and None != cv:
            return cv
        else:
            return ''

        pass


    def CellWrite(self, sht, value, startrow=0, startcol=0, endrow=0, endcol=0, style=None):
        if 'w' != self.option:
            LOG.warn("[Excel2007 CellWrite] not write option which is initted.")
            return

        if None == value:
            LOG.error("[Excel2007 CellWrite] input value is empty.")

        if endrow <= startrow:
            endrow = startrow
        if endcol <= startcol:
            endcol = startcol

        if (endrow == startrow) and (endcol == startcol):

            sht.cell(row=startrow, col=startcol, value=value)
        else:
            sht.cell(row=startrow, col=startcol, value=value)
            LOG.warn("[Excel2007 CellWrite] no merge function.")


        LOG.debug("range[(%d, %d)->(%d, %d)] write context %s." % (startrow, startcol, endrow, endcol, value))
        pass


    def SaveWriteWb(self):
        self.writewb.save(self.xlspath)
